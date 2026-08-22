"""Diagnostica: confronta per ogni giocatore il V.A. dell'Excel (colonna K)
con il valore ricalcolato VI×QA/QI arrotondato a 1 e 2 decimali."""
import sys

sys.path.insert(0, __file__.rsplit("\\", 1)[0])
from import_rose import EXCEL_PATH, TEAM_SHEETS, build_payload  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


def main() -> None:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    for sheet in TEAM_SHEETS:
        ws = wb[sheet]
        mismatches = []
        for row in range(5, ws.max_row + 1):
            name = ws.cell(row=row, column=3).value
            if name is None or str(name).strip() in ("", "-"):
                continue
            vi = float(ws.cell(row=row, column=10).value or 0)
            qi = float(ws.cell(row=row, column=8).value or 0)
            qa = float(ws.cell(row=row, column=9).value or 0)
            k_excel = float(ws.cell(row=row, column=11).value or 0)
            if not qi:
                continue
            raw = vi * qa / qi
            r1 = round(raw * 10 + (0.5 if raw >= 0 else -0.5)) / 10
            r2 = round(raw * 100 + (0.5 if raw >= 0 else -0.5)) / 100
            if abs(k_excel - r1) > 0.001 and abs(k_excel - r2) > 0.001:
                mismatches.append((name, vi, qi, qa, k_excel, raw, r1, r2))
        if mismatches:
            print(f"\n=== {sheet} ===")
            for m in mismatches:
                print(
                    f"  {m[0]}: VI={m[1]} QI={m[2]} QA={m[3]} | "
                    f"Excel K={m[4]} | raw={m[5]:.4f} | r1={m[6]} r2={m[7]}"
                )


if __name__ == "__main__":
    main()