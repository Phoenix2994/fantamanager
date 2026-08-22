"""
Import iniziale di ROSE.xlsx su Firestore.

Uso:
    py -3 scripts/import_rose.py            # DRY-RUN: solo parsing e riepilogo
    py -3 scripts/import_rose.py --write    # Import reale su Firestore

Requisiti per l'import reale:
    - py -3 -m pip install firebase-admin
    - Chiave service account salvata in scripts/serviceAccountKey.json
      (Console Firebase → Impostazioni progetto → Account di servizio →
       Genera nuova chiave privata → salva il JSON con quel nome)

Compatibile con il piano gratuito Spark: usa solo scritture Firestore
(~400 documenti, ben sotto il limite giornaliero gratuito di 20k).
Le security rules NON si applicano all'Admin SDK.

Struttura scritta (isolamento per stagione):
    league/{leagueId}
    league/{leagueId}/taxBrackets/{1..6}
    teams/{teamId}                                  (dati anagrafici squadra)
    teams/{teamId}/seasons/{season}/players/{playerId}
    teams/{teamId}/seasonFinance/{season}
    auditLog/{autoId}                               (voce initial_import)
"""
import argparse
import math
import os
import re
import sys
import unicodedata

from openpyxl import load_workbook

# ---------------------------------------------------------------- costanti
EXCEL_PATH = (
    r"C:\Users\filip\Dropbox\Fantacalcio Ufficiale"
    r"\Fantacalcio Ufficiale 2026-27\ROSE.xlsx"
)
KEY_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "serviceAccountKey.json")

LEAGUE_ID = "main"
LEAGUE_NAME = "Fantacalcio Manageriale 2026-27"
SEASON = "2026-27"

# I 10 fogli ufficiali della stagione corrente (esclusi i legacy
# "LEI CE STA FC" e "Vitanews", stagione 2019/20)
TEAM_SHEETS = [
    "Ac. Ciaccati",
    "S.S. Jonica 106",
    "Phoenix",
    "Cispo's Vision",
    "Granchi Avatori",
    "Nicaragua Pacamara Gigante",
    "Akatsuki",
    "Loco Barurumon",
    "DYNAMO COCITO",
    "DAS HAUS",
]

CONTRACT_DEFAULT = "TITOLO DEFINITIVO"


# ---------------------------------------------------------------- utilità
def slugify(text: str) -> str:
    """ID deterministico: ascii-fold, minuscole, non-alfanumerici → '-'."""
    normalized = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^a-z0-9]+", "-", normalized.lower()).strip("-")
    return slug


def round2(x: float) -> float:
    """Arrotonda a 2 decimali come Math.round di JavaScript."""
    return math.floor(x * 100 + 0.5) / 100 if x >= 0 else -math.floor(-x * 100 + 0.5) / 100


def num(value) -> float:
    """Converte una cella Excel in float (0.0 se vuota/non numerica)."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return 0.0


def clean_text(value) -> str:
    return str(value).strip() if value is not None else ""


def norm_label(value) -> str:
    """Normalizza un'etichetta: minuscole, spazi collassati, ':' finale via."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower().rstrip(":").strip()


def norm_contract(value) -> str:
    if value is None:
        return CONTRACT_DEFAULT
    text = re.sub(r"\s+", " ", str(value)).strip().upper()
    return text or CONTRACT_DEFAULT


# ---------------------------------------------------------------- parsing
def parse_tax_brackets(wb):
    """Scaglioni fiscali dal foglio 'Rose': AM5-AO10."""
    ws = wb["Rose"]
    brackets = []
    for i, row in enumerate(range(5, 11), start=1):
        aliquota = num(ws.cell(row=row, column=39).value)   # AM
        limite_perc = num(ws.cell(row=row, column=40).value)  # AN (info extra)
        soglia = num(ws.cell(row=row, column=41).value)     # AO
        brackets.append(
            {
                "bracketIndex": i,
                "aliquota": aliquota,
                "limitePercRinnovo": limite_perc,
                "limiteSogliaEuro": soglia,
            }
        )
    return brackets


def parse_players(ws):
    """Giocatori dalle righe 5+ del foglio squadra (colonne B-K)."""
    players = []
    seen_names = {}
    for row in range(5, ws.max_row + 1):
        name = clean_text(ws.cell(row=row, column=3).value)  # C
        if not name or name == "-":
            continue

        ruolo = clean_text(ws.cell(row=row, column=2).value)  # B
        contract = norm_contract(ws.cell(row=row, column=5).value)  # E
        perc = num(ws.cell(row=row, column=6).value)          # F
        qi = num(ws.cell(row=row, column=8).value)            # H
        qa = num(ws.cell(row=row, column=9).value)            # I
        vi = num(ws.cell(row=row, column=10).value)           # J
        spent = num(ws.cell(row=row, column=4).value)         # D (vuoto nella stagione corrente)

        # Fiducia ai valori ufficiali del foglio (K = V.A., G = spesa rinnovo):
        # l'Excel non applica una formula di arrotondamento consistente, quindi
        # li importiamo così come sono. Le formule dell'app (finance-calculator.ts)
        # entreranno in gioco solo per gli aggiornamenti futuri delle quotazioni.
        valore_attuale = round2(num(ws.cell(row=row, column=11).value))  # K
        prossima_spesa = round2(num(ws.cell(row=row, column=7).value))   # G

        base_slug = slugify(name)
        count = seen_names.get(base_slug, 0)
        seen_names[base_slug] = count + 1
        doc_id = base_slug if count == 0 else f"{base_slug}-{count + 1}"

        players.append(
            {
                "id": doc_id,
                "name": name,
                "ruolo": ruolo,
                "contractType": contract,
                "acquistoRinnovoSpesa": round2(spent),
                "prossimaPercRinnovo": perc,
                "prossimaSpesaRinnovo": prossima_spesa,
                "quotazioneIniziale": qi,
                "quotazioneAttuale": qa,
                "valoreIniziale": vi,
                "valoreAttuale": valore_attuale,
            }
        )
    return players


FIN_SIMPLE = {
    "rinnovi": "rinnovi",
    "acquisti all'asta": "acquistiAstaSettembre",
    "rescissioni": "rescissioni",
    "penali": "penali",
    "soldi versati": "soldiVersati",
    "premi": "premi",
    "rimborsi": "rimborsi",
    "indennizzi": "indennizzoSettembre",
}


def parse_finance(ws):
    """
    Spese societarie dalle colonne M/N.
    - 'Trasferimenti' compare due volte: 1ª occorrenza = uscita, 2ª = entrata.
    - I campi calcolati dell'Excel (tasse, spese, bilanci) vengono ignorati:
      sono ricalcolati con le formule dell'app.
    - 'Bilancio societario totale' è il patrimonio storico da preservare.
    """
    inputs = {
        "rinnovi": 0.0,
        "acquistiMercatoInfrasettimanale": 0.0,
        "acquistiAstaSettembre": 0.0,
        "acquistiAstaGennaio": 0.0,
        "rescissioni": 0.0,
        "penali": 0.0,
        "trasferimentiUscita": 0.0,
        "trasferimentiEntrata": 0.0,
        "indennizzoSettembre": 0.0,
        "indennizzoGennaio": 0.0,
        "rimborsi": 0.0,
        "premi": 0.0,
        "soldiVersati": 0.0,
    }
    bilancio_totale_precedente = 0.0
    trasferimenti_seen = 0

    for row in range(4, min(ws.max_row, 35) + 1):
        label = norm_label(ws.cell(row=row, column=13).value)  # M
        value = num(ws.cell(row=row, column=14).value)         # N
        if not label:
            continue

        if label == "trasferimenti":
            trasferimenti_seen += 1
            if trasferimenti_seen == 1:
                inputs["trasferimentiUscita"] = value
            else:
                inputs["trasferimentiEntrata"] = value
        elif label == "bilancio societario totale":
            bilancio_totale_precedente = value
        elif label in FIN_SIMPLE:
            inputs[FIN_SIMPLE[label]] = value
        # tutto il resto (tasse, spesa da versare, valore rosa, ...) è ricalcolato

    return inputs, round2(bilancio_totale_precedente)


def compute_finance(inputs, brackets, valore_rosa, historic=0.0):
    """Replica esatta delle formule di src/app/core/finance-calculator.ts."""
    spesa_annuale = round2(
        inputs["rinnovi"]
        + inputs["acquistiMercatoInfrasettimanale"]
        + inputs["acquistiAstaSettembre"]
        + inputs["acquistiAstaGennaio"]
        + inputs["rescissioni"]
        + inputs["penali"]
        + inputs["trasferimentiUscita"]
        - inputs["trasferimentiEntrata"]
        - inputs["rimborsi"]
    )

    ordered = sorted(brackets, key=lambda b: b["bracketIndex"])
    totale = 0.0
    for i, bracket in enumerate(ordered):
        low = bracket["limiteSogliaEuro"]
        high = ordered[i + 1]["limiteSogliaEuro"] if i + 1 < len(ordered) else math.inf
        if spesa_annuale > low:
            totale += (min(spesa_annuale, high) - low) * bracket["aliquota"]
    tassa_calcolata = round2(totale)

    tasse = max(tassa_calcolata, historic)
    tax_minimum_historic = max(tasse, historic)

    spesa_da_versare = round2(
        inputs["rinnovi"]
        + inputs["acquistiMercatoInfrasettimanale"]
        + max(0.0, inputs["acquistiAstaSettembre"] - inputs["indennizzoSettembre"])
        + max(0.0, inputs["acquistiAstaGennaio"] - inputs["indennizzoGennaio"])
        + inputs["rescissioni"]
        + inputs["penali"]
        - inputs["rimborsi"]
        + tasse
    )
    spesa_totale = round2(spesa_da_versare + inputs["trasferimentiUscita"])
    soldi_da_versare = round2(spesa_da_versare - inputs["soldiVersati"])
    bilancio_stagionale = round2(
        inputs["premi"] + inputs["trasferimentiEntrata"] - spesa_totale
    )

    return {
        "spesaAnnuale": spesa_annuale,
        "tasse": tasse,
        "taxMinimumHistoric": round2(tax_minimum_historic),
        "spesaDaVersare": spesa_da_versare,
        "spesaTotale": spesa_totale,
        "soldiDaVersare": soldi_da_versare,
        "valoreRosa": round2(valore_rosa),
        "bilancioSocietarioStagionale": bilancio_stagionale,
    }


def build_payload():
    """Legge l'Excel e costruisce tutti i payload (nessuna scrittura)."""
    wb = load_workbook(EXCEL_PATH, data_only=True)
    brackets = parse_tax_brackets(wb)

    teams = []
    for sheet_name in TEAM_SHEETS:
        ws = wb[sheet_name]
        team_name = clean_text(ws.cell(row=2, column=2).value) or sheet_name
        players = parse_players(ws)
        inputs, bilancio_prec = parse_finance(ws)

        valore_rosa = round2(sum(p["valoreAttuale"] for p in players))
        excel_valore_rosa = num(ws.cell(row=20, column=14).value)  # N20 (validazione)
        computed = compute_finance(inputs, brackets, valore_rosa)
        bilancio_totale = round2(bilancio_prec + computed["bilancioSocietarioStagionale"])

        teams.append(
            {
                "id": slugify(team_name),
                "name": team_name,
                "sheet": sheet_name,
                "players": players,
                "financeInputs": inputs,
                "financeComputed": computed,
                "bilancioTotalePrecedente": bilancio_prec,
                "bilancioTotale": bilancio_totale,
                "excelValoreRosa": round2(excel_valore_rosa),
            }
        )
    return brackets, teams


# ---------------------------------------------------------------- scrittura
def write_to_firestore(brackets, teams):
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore
    except ImportError:
        sys.exit("ERRORE: installa firebase-admin →  py -3 -m pip install firebase-admin")

    if not os.path.exists(KEY_PATH):
        sys.exit(
            "ERRORE: manca scripts/serviceAccountKey.json.\n"
            "Console Firebase → Impostazioni progetto → Account di servizio →\n"
            "Genera nuova chiave privata → salva il JSON come "
            "scripts/serviceAccountKey.json"
        )

    cred = credentials.Certificate(KEY_PATH)
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    ops = 0
    batches_committed = 0
    batch = db.batch()

    def add(ref, data):
        nonlocal batch, ops, batches_committed
        batch.set(ref, data)
        ops += 1
        if ops >= 450:
            batch.commit()
            batches_committed += 1
            batch = db.batch()
            ops = 0

    # 1. Lega
    add(db.collection("league").document(LEAGUE_ID),
        {"name": LEAGUE_NAME, "season": SEASON})

    # 2. Scaglioni fiscali
    for bracket in brackets:
        add(db.collection("league").document(LEAGUE_ID)
            .collection("taxBrackets").document(str(bracket["bracketIndex"])),
            bracket)

    # 3. Squadre + rose + finanze di stagione
    for team in teams:
        team_ref = db.collection("teams").document(team["id"])
        add(team_ref, {
            "leagueId": LEAGUE_ID,
            "name": team["name"],
            "bilancioTotalePrecedente": team["bilancioTotalePrecedente"],
            "bilancioTotale": team["bilancioTotale"],
        })

        season_ref = team_ref.collection("seasons").document(SEASON)
        for player in team["players"]:
            add(season_ref.collection("players").document(player["id"]), player)

        add(team_ref.collection("seasonFinance").document(SEASON), {
            **team["financeInputs"],
            **team["financeComputed"],
        })

    # 4. Voce audit log per l'import iniziale
    add(db.collection("auditLog").document(), {
        "timestamp": firestore.SERVER_TIMESTAMP,
        "leagueId": LEAGUE_ID,
        "teamId": "",
        "adminId": "import-script",
        "entityType": "initial_import",
        "entityId": SEASON,
        "operation": "create",
        "fieldModified": "*",
        "valueBefore": None,
        "valueAfter": len(teams),
        "changeSummary": f"Import iniziale ROSE.xlsx — stagione {SEASON}",
    })

    if ops > 0:
        batch.commit()
        batches_committed += 1

    total_docs = 1 + len(brackets) + sum(
        1 + len(t["players"]) + 1 for t in teams
    ) + 1
    print(f"Import completato: ~{total_docs} documenti in {batches_committed} batch.")


# ---------------------------------------------------------------- main
def main():
    parser = argparse.ArgumentParser(description="Import ROSE.xlsx → Firestore")
    parser.add_argument("--write", action="store_true",
                        help="Esegue davvero la scrittura su Firestore "
                             "(default: solo dry-run)")
    args = parser.parse_args()

    print(f"Lettura di: {EXCEL_PATH}\n")
    brackets, teams = build_payload()

    print(f"SCAGLIONI FISCALI ({len(brackets)}):")
    for b in brackets:
        print(f"  #{b['bracketIndex']}: aliquota={b['aliquota']} "
              f"soglia={b['limiteSogliaEuro']} €")

    tot_players = 0
    print("\nSQUADRE:")
    for team in teams:
        tot_players += len(team["players"])
        fc = team["financeComputed"]
        diff_rosa = abs(fc["valoreRosa"] - team["excelValoreRosa"])
        warn = "" if diff_rosa < 0.05 else \
            f"  ⚠ valore rosa diverso dall'Excel ({team['excelValoreRosa']})"
        print(f"\n  ▶ {team['name']} (id: {team['id']}) — {len(team['players'])} giocatori{warn}")
        print(f"    Valore rosa calcolato: {fc['valoreRosa']} € "
              f"(Excel: {team['excelValoreRosa']} €)")
        print(f"    Inputs: rinnovi={team['financeInputs']['rinnovi']} "
              f"astaSett={team['financeInputs']['acquistiAstaSettembre']} "
              f"rescissioni={team['financeInputs']['rescissioni']} "
              f"penali={team['financeInputs']['penali']} "
              f"trasfUscita={team['financeInputs']['trasferimentiUscita']} "
              f"trasfEntrata={team['financeInputs']['trasferimentiEntrata']} "
              f"premi={team['financeInputs']['premi']} "
              f"rimborsi={team['financeInputs']['rimborsi']} "
              f"indennizzi={team['financeInputs']['indennizzoSettembre']} "
              f"soldiVersati={team['financeInputs']['soldiVersati']}")
        print(f"    Calcolati: spesaAnnuale={fc['spesaAnnuale']} "
              f"tasse={fc['tasse']} spesaDaVersare={fc['spesaDaVersare']} "
              f"spesaTotale={fc['spesaTotale']} "
              f"soldiDaVersare={fc['soldiDaVersare']} "
              f"bilancioStagionale={fc['bilancioSocietarioStagionale']}")
        print(f"    Bilancio totale: precedente={team['bilancioTotalePrecedente']} "
              f"→ {team['bilancioTotale']} €")

    print(f"\nTOTALE: {len(teams)} squadre, {tot_players} giocatori, "
          f"{len(brackets)} scaglioni")

    if args.write:
        print("\n=== SCRITTURA SU FIRESTORE IN CORSO ===")
        write_to_firestore(brackets, teams)
    else:
        print("\nDRY-RUN completato: nessuna scrittura eseguita.")
        print("Per importare davvero:  py -3 scripts/import_rose.py --write")


if __name__ == "__main__":
    main()