"""Ispezione read-only di ROSE.xlsx: stampa nomi fogli e celle non vuote
delle prime righe, per capire la struttura prima dell'import su Firestore."""
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\filip\Dropbox\Fantacalcio Ufficiale\Fantacalcio Ufficiale 2026-27\ROSE.xlsx"
MAX_ROWS = int(sys.argv[1]) if len(sys.argv) > 1 else 40
MAX_COLS = int(sys.argv[2]) if len(sys.argv) > 2 else 60


def main() -> None:
    wb = load_workbook(PATH, data_only=True)
    print("SHEETS:", wb.sheetnames)

    for name in wb.sheetnames:
        ws = wb[name]
        print(
            f"\n===== SHEET: {name} | max_row={ws.max_row} "
            f"max_col={ws.max_column} ====="
        )
        max_r = min(ws.max_row or 0, MAX_ROWS)
        max_c = min(ws.max_column or 0, MAX_COLS)
        for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
            cells = []
            for cell in row:
                if cell.value is not None:
                    col = get_column_letter(cell.column)
                    value = str(cell.value).replace("\n", "\\n")
                    if len(value) > 40:
                        value = value[:37] + "..."
                    cells.append(f"{col}{cell.row}={value}")
            if cells:
                print(" | ".join(cells))


if __name__ == "__main__":
    main()